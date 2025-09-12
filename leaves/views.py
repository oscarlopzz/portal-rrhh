from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import LeaveRequest
from .forms import LeaveRequestCreateForm

from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils.dateparse import parse_date
import csv

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

@login_required
def my_leaves(request):
    employee = getattr(request.user, 'employee_profile', None)
    if not employee:
        messages.error(request, "Tu usuario no está vinculado a un empleado.")
        return redirect('home')

    status = request.GET.get('status') or ''
    from_str = request.GET.get('from') or ''
    to_str = request.GET.get('to') or ''
    d_from = parse_date(from_str) if from_str else None
    d_to = parse_date(to_str) if to_str else None

    qs = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')

    if status:
        qs = qs.filter(status=status)
    if d_from:
        qs = qs.filter(start_date__gte=d_from)
    if d_to:
        qs = qs.filter(end_date__lte=d_to)

    return render(request, 'leaves/my_list.html', {
        'items': qs,
        'status': status,
        'from': from_str,
        'to': to_str,
    })


def home(request):
    # /leaves/ -> redirige a /leaves/mis/
    return redirect('my_leaves')

@login_required
def create_leave(request):
    employee = getattr(request.user, 'employee_profile', None)
    if not employee:
        messages.error(request, "Tu usuario no está vinculado a un empleado.")
        return redirect('home')

    if request.method == 'POST':
        form = LeaveRequestCreateForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.status = LeaveRequest.PENDING  # por defecto
            leave.save()
            messages.success(request, "Solicitud de ausencia creada y pendiente de aprobación.")
            return redirect('my_leaves')
    else:
        form = LeaveRequestCreateForm()

    return render(request, 'leaves/create.html', {'form': form})

# (Opcionales, si ya los añadiste para RRHH/Supervisor)
@login_required
@permission_required('leaves.can_approve_leaves', raise_exception=True)
def review_list(request):
    status = request.GET.get('status') or ''
    q = request.GET.get('q') or ''
    from_str = request.GET.get('from') or ''
    to_str = request.GET.get('to') or ''
    d_from = parse_date(from_str) if from_str else None
    d_to = parse_date(to_str) if to_str else None

    qs = LeaveRequest.objects.select_related('employee').order_by('-created_at')

    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q) |
            Q(employee__last_name__icontains=q) |
            Q(employee__email__icontains=q)
        )
    if d_from:
        qs = qs.filter(start_date__gte=d_from)
    if d_to:
        qs = qs.filter(end_date__lte=d_to)

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'leaves/review_list.html', {
        'items': page_obj,
        'status': status, 'q': q, 'from': from_str, 'to': to_str,
    })

@login_required
@permission_required('leaves.can_approve_leaves', raise_exception=True)
def review_export_csv(request):
    status = request.GET.get('status') or ''
    q = request.GET.get('q') or ''
    d_from = parse_date(request.GET.get('from') or '')
    d_to = parse_date(request.GET.get('to') or '')
    sep = (request.GET.get('sep') or 'comma').lower()

    delimiter = {'comma': ',', 'semicolon': ';', 'tab': '\t'}.get(sep, ',')
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp.write('\ufeff')
    resp['Content-Disposition'] = 'attachment; filename="ausencias_revisar.csv"'

    qs = LeaveRequest.objects.select_related('employee').order_by('-created_at')
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q) |
            Q(employee__last_name__icontains=q) |
            Q(employee__email__icontains=q)
        )
    if d_from:
        qs = qs.filter(start_date__gte=d_from)
    if d_to:
        qs = qs.filter(end_date__lte=d_to)

    w = csv.writer(resp, delimiter=delimiter)
    w.writerow(['ID','Empleado','Email','Tipo','Inicio','Fin','Días','Estado','Justificada','Motivo'])

    for l in qs:
        dias = (l.end_date - l.start_date).days + 1 if l.start_date and l.end_date else ''
        w.writerow([
            l.id,
            str(l.employee), getattr(l.employee, 'email', ''),
            l.get_ltype_display(), l.start_date, l.end_date, dias,
            l.get_status_display(), 'Sí' if l.is_justified else 'No', l.reason or ''
        ])
    return resp


@login_required
@permission_required('leaves.can_approve_leaves', raise_exception=True)
def review_detail(request, pk):
    item = get_object_or_404(LeaveRequest, pk=pk)
    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision in (LeaveRequest.APPROVED, LeaveRequest.REJECTED):
            item.status = decision
            item.save()
            messages.success(request, "Decisión registrada.")
            return redirect('leaves_review_list')
        messages.error(request, "Decisión no válida.")
    return render(request, 'leaves/review_detail.html', {'item': item})



@login_required
def my_leaves_export_csv(request):
    employee = getattr(request.user, 'employee_profile', None)
    if not employee:
        return HttpResponse(status=403)

    status = request.GET.get('status') or ''
    d_from = parse_date(request.GET.get('from') or '')
    d_to = parse_date(request.GET.get('to') or '')
    sep = (request.GET.get('sep') or 'comma').lower()   # comma | semicolon | tab

    # Elige el delimitador
    delimiter = {'comma': ',', 'semicolon': ';', 'tab': '\t'}.get(sep, ',')
    # BOM para que Excel detecte UTF-8
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp.write('\ufeff')
    resp['Content-Disposition'] = 'attachment; filename="mis_ausencias.csv"'

    qs = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')
    if status:
        qs = qs.filter(status=status)
    if d_from:
        qs = qs.filter(start_date__gte=d_from)
    if d_to:
        qs = qs.filter(end_date__lte=d_to)

    w = csv.writer(resp, delimiter=delimiter)
    w.writerow(['ID', 'Tipo', 'Inicio', 'Fin', 'Días', 'Estado', 'Justificada', 'Motivo'])

    for l in qs:
        dias = (l.end_date - l.start_date).days + 1 if l.start_date and l.end_date else ''
        w.writerow([
            l.id,
            l.get_ltype_display(),
            l.start_date,    # Excel/LO los detecta como fecha
            l.end_date,
            dias,
            l.get_status_display(),
            'Sí' if l.is_justified else 'No',
            l.reason or ''
        ])
    return resp

@login_required
def my_leaves_export_xlsx(request):
    employee = getattr(request.user, 'employee_profile', None)
    if not employee:
        return HttpResponse(status=403)

    status = request.GET.get('status') or ''
    d_from = parse_date(request.GET.get('from') or '')
    d_to   = parse_date(request.GET.get('to') or '')

    qs = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')
    if status: qs = qs.filter(status=status)
    if d_from: qs = qs.filter(start_date__gte=d_from)
    if d_to:   qs = qs.filter(end_date__lte=d_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mis ausencias"
    headers = ['ID','Tipo','Inicio','Fin','Días','Estado','Justificada','Motivo']
    ws.append(headers)

    for l in qs:
        dias = (l.end_date - l.start_date).days + 1 if l.start_date and l.end_date else ''
        ws.append([
            l.id, l.get_ltype_display(), l.start_date, l.end_date, dias,
            l.get_status_display(), 'Sí' if l.is_justified else 'No', l.reason or ''
        ])

    # Estilos: encabezado en negrita, centrado y congelar fila 1
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = "A2"

    # Formato de fecha (columnas C y D) y ancho automático
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for c in row:
            c.number_format = 'yyyy-mm-dd'

    for col in range(1, ws.max_column + 1):
        length = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, length + 2), 40)

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="mis_ausencias.xlsx"'
    wb.save(resp)
    return resp

from django.contrib.auth.decorators import login_required, permission_required

@login_required
@permission_required('leaves.can_approve_leaves', raise_exception=True)
def review_export_xlsx(request):
    status = request.GET.get('status') or ''
    q      = request.GET.get('q') or ''
    d_from = parse_date(request.GET.get('from') or '')
    d_to   = parse_date(request.GET.get('to') or '')

    qs = LeaveRequest.objects.select_related('employee').order_by('-created_at')
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q) |
            Q(employee__last_name__icontains=q) |
            Q(employee__email__icontains=q)
        )
    if d_from:
        qs = qs.filter(start_date__gte=d_from)
    if d_to:
        qs = qs.filter(end_date__lte=d_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Revisión de ausencias"
    headers = ['ID','Empleado','Email','Tipo','Inicio','Fin','Días','Estado','Justificada','Motivo']
    ws.append(headers)

    for l in qs:
        dias = (l.end_date - l.start_date).days + 1 if l.start_date and l.end_date else ''
        ws.append([
            l.id, str(l.employee), getattr(l.employee, 'email', ''),
            l.get_ltype_display(), l.start_date, l.end_date, dias,
            l.get_status_display(), 'Sí' if l.is_justified else 'No', l.reason or ''
        ])

    # Estilos
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = "A2"

    # Formato fecha (Inicio/Fin) + anchos
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for c in row:
            c.number_format = 'yyyy-mm-dd'
    for col in range(1, ws.max_column + 1):
        length = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, length + 2), 40)

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="ausencias_revisar.xlsx"'
    wb.save(resp)
    return resp

